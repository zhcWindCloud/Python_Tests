import re
import warnings

import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

ua = UserAgent()

headers = {'User-Agent': ua.random}

proxies = {'http': 'http://187.49.83.153:8080', 'https': 'https://187.49.83.153:8080'}


def GetHtml(url):
    try:
        html = requests.get(url, headers=headers, timeout=5,  verify=False)
        html.encoding = "utf-8"
        if html.status_code == 200:
            bs = BeautifulSoup(html.content, 'lxml')
            return bs
    except Exception as e:
        print(e)
        return None


def GetArtist(bs):
    """获取歌手"""
    if bs:
        artist = {}
        soups = bs.findAll("a",
                           attrs={"href": re.compile("artist"), "title": re.compile("\w+"), "class": re.compile("nm")})
        for soup in soups:
            if soup.get_text():
                artist[soup.get_text()] = "https://music.163.com{0}".format(soup.get("href").strip())
        return artist
    return "request error"


def GetArtistSong(artists):
    """获取歌手歌曲"""
    Songs = {}
    if isinstance(artists,dict):
        for artist, songUrl in artists.items():
            bs = GetHtml(songUrl)
            if bs:
                soups = bs.findAll("a", attrs={"href": re.compile("song.id=(\d+)")})
                songs = {}
                for soup in soups:
                    songs[soup.get_text()] = re.compile('\d+').findall(soup.get("href"))[0]
                Songs[artist] = songs
        return Songs


def GetDownload(SongsDict):
    """下载歌曲，并保存到本地工作路径
    :param SongsDict 歌曲字典信息
    """
    if isinstance(SongsDict,dict):
        for artist, songInfo in SongsDict.items():
            root = r'F:\SpiderMusic\{0}'.format(artist)
            mkdir(root)
            for songName, songID in songInfo.items():
                print("http://music.163.com/song/media/outer/url?id=" + songID)
                path = root + r'\{0}.mp3'.format(songName)
                try:
                    r = requests.get("http://music.163.com/song/media/outer/url?id=" + songID, headers=headers, timeout=5)
                    print(r.status_code, r.content)
                    print('{0}下载成功'.format(songName))
                except Exception as e:
                    print(e)




def Download():
    """从excel中找到歌曲ID，并保存到本地工作路径
    :param SongsDict 歌曲字典信息
    """
    excel_adress = r"F:\Python爬虫测试.xlsx"
    wb = load_workbook(excel_adress)
    sheet = wb[wb.sheetnames[0]]
    # for row in sheet.rows:
    #     for cell in row:
    #         path =r'F:\SpiderMusic\{0}.mp3'.format(cell.value)
    #         r = requests.get("http://music.163.com/song/media/outer/url?id=" + songID)
    #         with open(path, "wb") as file:
    #             file.write(r.content)


def SaveSongInfo(SongsDict):
    """保存歌曲信息"""
    excel_adress = r"F:\Python爬虫测试.xlsx"
    wb = load_workbook(excel_adress)
    sht = wb[wb.sheetnames[0]]
    for artist, songInfo in SongsDict.items():
        for songName, songID in songInfo.items():
            songurl = "http://music.163.com/song/media/outer/url?id={0}".format(songID)
            songlist = [songName, songurl]
            sht.append(songlist)
    wb.save(excel_adress)
    print("{0}保存成功！".format(excel_adress))


def IS_ip(IP, Port):
    """判断IP代理是否有效"""
    import telnetlib

    try:
        telnetlib.Telnet(IP, port=Port, timeout=3)
    except:
        return False
    else:
        return True


def mkdir(path):
    """创建目录"""
    # 引入模块
    import os

    # 去除首位空格
    path = path.strip()
    # 去除尾部 \ 符号
    path = path.rstrip("\\")

    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    isExists = os.path.exists(path)

    # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        # 创建目录操作函数
        os.makedirs(path)

        print(path + ' 创建成功')
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(path + ' 目录已存在')
        return False


if __name__ == '__main__':
    import time

    start = time.time()
    ArtistNo = [1001, 1002, 1003]
    for x in ArtistNo:
        url = "https://music.163.com/discover/artist/cat?id={0}".format(str(x))
        bs = GetHtml(url)

        artists = GetArtist(bs)

        SonsDict = GetArtistSong(artists)

        GetDownload(SonsDict)
    end = time.time()
    print("总共花费{0}秒".format(end - start))
