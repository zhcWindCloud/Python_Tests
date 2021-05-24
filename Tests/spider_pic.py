import time
import warnings

import openpyxl
import requests
from bs4 import BeautifulSoup

warnings.filterwarnings("ignore")
from fake_useragent import UserAgent

time_start = time.time()

ua = UserAgent(use_cache_server=False)

print('随机打印 User-Agent:', ua.random, type(ua.random))  ## 随机打印 User-Agent

print('随机打印一个 ie 浏览器的头', ua.ie)  # 随机打印一个 ie 浏览器的头

# 爬取登录后的超星学习通
# headers = {# 超星学习通的cookie
#           'Cookie': 'isfyportal=1; uname=2017207325055; lv=3; fid=2076; _uid=60597395; uf=da0883eb5260151e235da448df540e8d20d72176826470bd47f47e7dfbd23aa4a9c0fbadfefa8b0cba4fe937fc5d4fc0c49d67c0c30ca5047c5a963e85f110997d251c7e343296a4ce71fc6e59483dd35c1b59c70670ec645ba9852d59d9559bcafeb12deae5fab7; _d=1603026163348; UID=60597395; vc=D5A85263598D2D440007C305BBF43D24; vc2=997BBAF3CF94A264B7BBD75FCE94614A; vc3=ctfY2G1MxWJRpAjaI8IBsveyU%2BMPQYunXG3CH3CCk4YcP2ZzeQYpnYz0v%2BEPLVGMgrjUVaDWG%2B%2BNw0Z89lk3fNtig98LBuoRYhrxSKbfvPqlIsTLUBrveHdWguvSRAATCCDeBLRwxAIx7OX25CeA8QYv83GAf4LN92IYShzEDeY%3D48c1c0bb9395ce7cfd695792f93a3282; xxtenc=8a123585b993768d5122b8f82fb86d8a; DSSTASH_LOG=C_38-UN_525-US_60597395-T_1603026163350; fanyamoocs=975B3FB9EE89B0EF11401F839C536D9E; JSESSIONID=783B95175A7D5039D8BEF8EE65A4F7C1; route=98f17f0b052079b916f022bfaec8cf55; thirdRegist=0; tl=1',
#            'User-Agent': ua.random
# }
#
# url = "http://i.mooc.chaoxing.com/space/index.shtml"  #超星学习通的个人主页
#
# url_student = "http://mooc1-1.chaoxing.com/visit/interaction?s=f0e066bffa6d30f2c8f707576f8f66d4" #超星学习通的学生主页
#
#
# html = requests.get(url_student, headers=headers)  # 请求方式
#
# html.encoding ="utf-8"
#
# bs1 = BeautifulSoup(html.content)
#
#
# # print(bs1.find_all(name="p",attrs={"title":re.compile('(\w+)')}))   #获取相对应的列表
#
#
# l = bs1.find_all(name="p",attrs={"title":re.compile('(\w+)')})
# for x in l :
#     print(x.get("title"))
page = ['22', '23', '62', '83', '102', '122', '142', '162']

#
# headers = {
#     "User_Agent": ua.random,
#     'cookie': 'semester.id=22; JSESSIONID=267ACF9D7511F7C9D6ACEF3FFE1D0062'
# }
#
# url = "http://jxzx.ahtcm.edu.cn/eams/teach/grade/course/person!search.action?semesterId=" + str(
#     22) + "&projectType="
#
# html = requests.get(url, headers=headers)
# html.encoding = "utf-8"
# bs1 = BeautifulSoup(html.content)
#
# soup_th = bs1.find_all(name='th')
# soup_td = bs1.find_all(name="td")
#
# th_list = []
# td_list = []
# for x in soup_th:
#     th_list.append(x.text)
#
# for x in soup_td:
#     # print(x.text.strip('\t\n').strip(" ").strip("\t"))
#     n = x.text.strip('\t\n').strip(" ").strip("\t")
#     td_list.append(n)
#
# list = []
# lenth = int(len(td_list) / len(th_list))
# for i in range(lenth):
#     list.append(td_list[i * 9:i * 9 + 9])
# # print(list)
#
# for i in range(len(list)):
#     print(list[i],'**'*41)
#     for x in range(len(th_list)):
#         print(list[i][x])
#
# 'cookie': 'semester.id=23; JSESSIONID=7019CF13B8A16427D741C86D0CD9D3CB'  #朱华超

work_book = openpyxl.Workbook()
work_sheet = work_book.create_sheet(title="成绩单")
work_sheet.append(['期末成绩'])


def getHtml(cookie):  # 爬取函数
    headers = {
        "User_Agent": ua.random,
        'cookie': cookie
    }
    url = "http://jxzx.ahtcm.edu.cn/eams/teach/grade/course/person!historyCourseGrade.action?projectType=MAJOR"

    try:
        html = requests.get(url, headers=headers)
        html.encoding = "utf-8"
        bs1 = BeautifulSoup(html.content,"lxml")

        soup_table = bs1.find_all(class_="gridtable")
    except Exception as e:
        print(e)

    return soup_table


def AngleData(cookie):
    titleList = []
    tables = getHtml(cookie)
    for table in tables:
        bs = BeautifulSoup(str(table),"lxml")
        soup_trs = bs.find_all(name="tr")
        for soup in soup_trs:
            bs1 = BeautifulSoup(str(soup), "lxml")
            soup_th = bs1.find_all(name="th")
            soup_td = bs1.find_all(name="td")










def Save_Grade(page, cookie):
    global max, mp
    thlist = getHtml(page, cookie)[0]
    list1 = getHtml(page, cookie)[1]
    lenth = getHtml(page, cookie)[2]
    print(len(list1), len(thlist))
    print(thlist)
    mp = len(thlist)

    for i in range(len(list1)):
        print(list1[i], '**' * 41)

    for row in range(len(list1)):
        if row == 0:
            work_sheet.append(thlist)
            # for col in range(len(thlist)):
            #     work_sheet.cell(row + 1, col + 1).value = thlist[col]
        else:
            work_sheet.append(list1[row - 1])
            # for col in range(len(thlist)):
            #     work_sheet.cell(row + 1, col + 1).value = list1[row - 1][col]

    work_book.save(r"E:\大学成绩单.xlsx")


if __name__ == '__main__':
    cookie = input("请输入你的Cookie值：")
    AngleData(cookie)
    print("完成")

# 爬取图片
# def spider_pic():
#     headers = {'User-Agent': random.choice(ua)}
#     title_dict ={"rili":"日历","dongman":"动漫","fengjing":"风景","meinv":"美女","tiyu":"体育","keai":"可爱"}
#
#
#     for title in title_dict:
#         Img_link = []
#         for number in range(1,10):
#             url = "http://www.netbian.com/"+str(title)+"/index_"+str(number+1)+".htm"
#             # url ="http://www.netbian.com/index_3.htm"
#
#             html = requests.get(url,headers=headers)
#
#             html.encoding="utf-8"
#
#             bs1 = BeautifulSoup(html.content)
#
#             img_list = bs1.find_all(name='img',attrs={"alt":re.compile('(\w)+')})
#             Img_link = img_list+Img_link
#
#
#
#
# time_end=time.time()
# print('totally cost',time_end-time_start)
