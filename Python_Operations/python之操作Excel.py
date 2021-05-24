from  openpyxl import Workbook,load_workbook


#写入
excel_adress = r"F:\Python操作Excel测试.xlsx"

wb = load_workbook(excel_adress)
sht = wb[wb.sheetnames[0]]


print()    ##sheetnames是获取工作表中的所有表的表名，即可查看该xlsx文件里面所有表

print(sht)


# 没有定位直接写入
list_1 = [5, 6, 7, 8]
sht.append(list_1)

# 增加一个定位
sht["A3"] = "开始"
list_2 = [1, 2, 3, 4]
sht.append(list_2)

# 增加一个定位
sht["D5"] = "新开始"
list_3 = [11, 12, 13, 14]
sht.append(list_3)

# 增加一个定位
sht["L15"] = "新开始2"
list_4 = [21, 22, 23, 24]
list_5 = [21, "赌气回答我的", 23, 24]
sht.append(list_4)
sht.append(list_5)

wb.save(excel_adress)






